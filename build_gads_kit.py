# -*- coding: utf-8 -*-
# Language rule: file is ENGLISH. Only assets that go LIVE into the German campaign
# (keywords, ad headlines/descriptions, callouts, sitelink text, snippet values) stay
# German — always with a separate English-translation column.
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY="1A1A2E"; RED="C8102E"; LIGHT="F2F2F5"; BAND="E9E9EF"; GREEN="1E7A34"; AMBER="B25C00"; WHITE="FFFFFF"; GREY="6B6B7B"
ARIAL="Arial"
thin=Side(style="thin",color="D5D5DD"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
def hfont(sz=11,bold=True,color=WHITE): return Font(name=ARIAL,size=sz,bold=bold,color=color)
def cfont(sz=10,bold=False,color="000000",italic=False): return Font(name=ARIAL,size=sz,bold=bold,color=color,italic=italic)
center=Alignment(horizontal="center",vertical="center")
left=Alignment(horizontal="left",vertical="center")
wrap=Alignment(horizontal="left",vertical="center",wrap_text=True)
topleft=Alignment(horizontal="left",vertical="top",wrap_text=True)

OVER=[]
def chk(text,limit,where):
    n=len(text)
    if n>limit: OVER.append((where,text,n,limit))
    return n,("OK" if n<=limit else f"! {n-limit} over")

def title_block(ws,title,subtitle,ncols):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
    c=ws.cell(1,1,title); c.font=hfont(16); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment(horizontal="left",vertical="center",indent=1); ws.row_dimensions[1].height=30
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncols)
    c=ws.cell(2,1,subtitle); c.font=cfont(9,italic=True,color=WHITE); c.fill=PatternFill("solid",fgColor=RED); c.alignment=Alignment(horizontal="left",vertical="center",indent=1); ws.row_dimensions[2].height=18
def header_row(ws,row,headers):
    for j,h in enumerate(headers,1):
        c=ws.cell(row,j,h); c.font=hfont(10); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=center; c.border=border
    ws.row_dimensions[row].height=30
def band(ws,row,text,ncols,color=BAND,fg=NAVY):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncols)
    c=ws.cell(row,1,text); c.font=cfont(10,bold=True,color=fg); c.fill=PatternFill("solid",fgColor=color); c.alignment=Alignment(horizontal="left",vertical="center",indent=1); ws.row_dimensions[row].height=20

wb=Workbook()

# ============================================================ 1. SETUP & METHOD
ws=wb.active; ws.title="1. Setup & Method"
title_block(ws,"DMR · Google Ads Search Campaign — Build Kit",
            "One ad group · Germany · $30/day · 1 month · Primary goal: Calendly booking (secondary: form). File is in English; live German assets carry a translation.",3)
ws.column_dimensions['A'].width=26; ws.column_dimensions['B'].width=66; ws.column_dimensions['C'].width=58
header_row(ws,4,["Setting","What to do","Why it matters"])
setup=[
 ("Campaign type","Search Network only. UNCHECK 'Include Google Display Network' and 'Search partners'.","They waste a $30/day budget on low-intent placements."),
 ("Goal","Leads — or 'Create a campaign without a goal' for full manual control.","Keeps every setting unrestricted."),
 ("Conversions","Mark BOTH Calendly booking AND form submit as PRIMARY.","On a tiny budget you need both to reach the ~15-conversion learning threshold."),
 ("Conversion values (rec.)","Give the booking a higher value than the form (e.g. 3 vs 1); bid on 'Maximize conversion value'.","Steers toward higher-value bookings without losing volume."),
 ("Location","Germany. Location options → 'Presence: people in your targeted locations' (NOT 'Presence or interest').","Stops foreign / irrelevant clicks burning budget."),
 ("Language","German.","The audience searches in German."),
 ("Daily budget","$30/day (≈ €27). Daily spend can spike up to 2×; the monthly average holds.","Plan for ~150–240 clicks/month total."),
 ("Bidding — weeks 1–2","'Maximize clicks' WITH a max CPC cap ≈ €3.50–€4.00.","Gathers keyword/click data fast without overpaying."),
 ("Bidding — week 3+","Switch to 'Maximize conversions' (optionally a target CPA) once you have ~15 conversions.","Enough signal for Smart Bidding to work."),
 ("Ad schedule","Launch ALL 7 days, ALL hours (at most trim 00:00–06:00). After ~3 weeks, use the 'When → Day & hour' report to bid up/down by time — don't hard-cut up front.","Calendly + form work 24/7, and busy SMB owners research evenings & weekends. No phone line = no reason to limit to office hours. Smart Bidding (week 3+) also auto-adjusts for time-of-day."),
 ("Devices","All on. Observe mobile first, adjust the bid later.","Don't pre-judge device performance."),
 ("Ad group","Exactly ONE: 'Google Ads — Agency & Management'. Keywords on tab 2.","Your specialty; concentrates the budget."),
 ("Keyword match","Phrase match for all keywords.","Broad match burns a small budget on loose variants."),
 ("Ads","Two responsive search ads (Set A + Set B) as an A/B test. Ad rotation: 'Optimize'.","Tests pain-point vs benefit messaging."),
 ("Extensions / assets","Add sitelinks, callouts and structured snippets (tabs 5–6).","Higher CTR and more ad space — free."),
 ("Negative keywords","Add the tab-3 list BEFORE launch.","Blocks job / course / free / DIY searchers."),
 ("Audiences","Add all as OBSERVATION (not Targeting) — tab 7.","Learn what converts without cutting reach."),
 ("Conversion tracking","Before launch confirm ONLY Calendly + Lead Form are Primary; set the 26 'BC' button-click events to Secondary.","Otherwise bidding chases button clicks, not real leads."),
]
r=5
for a,b,c in setup:
    ws.cell(r,1,a).font=cfont(10,bold=True,color=NAVY); ws.cell(r,1).alignment=topleft; ws.cell(r,1).border=border; ws.cell(r,1).fill=PatternFill("solid",fgColor=LIGHT)
    ws.cell(r,2,b).font=cfont(10); ws.cell(r,2).alignment=topleft; ws.cell(r,2).border=border
    ws.cell(r,3,c).font=cfont(9,color=GREY); ws.cell(r,3).alignment=topleft; ws.cell(r,3).border=border
    ws.row_dimensions[r].height=42; r+=1
r+=1
band(ws,r,"THE ANGLE",3); r+=1
angle=[
 ("Core message","The honest agency that fixes the leak: turn clicks into customers — measurable, transparent, no annual lock-in."),
 ("Why pain-point","Real German searches show frustration ('lots of clicks, no customers', 'budget burned'). Set B mirrors that; Set A leads with benefit/proof. Split-test both."),
 ("Conversion path","Ad → landing anchor (#kontakt / #pilot) → Calendly booking (primary) or form (secondary)."),
]
for a,b in angle:
    ws.cell(r,1,a).font=cfont(10,bold=True,color=RED); ws.cell(r,1).alignment=topleft; ws.cell(r,1).border=border; ws.cell(r,1).fill=PatternFill("solid",fgColor=LIGHT)
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
    ws.cell(r,2,b).font=cfont(10); ws.cell(r,2).alignment=topleft; ws.cell(r,2).border=border
    ws.cell(r,3).border=border
    ws.row_dimensions[r].height=44; r+=1
ws.freeze_panes="A5"

# ============================================================ 2. KEYWORDS (ranked pools)
ws=wb.create_sheet("2. Keywords")
title_block(ws,"Keywords — ONE ad group, ranked into search pools",
            "Same-meaning keywords are JOINED into one pool (they match the same searches). Priority Score /100 = how important the pool is (100 = use first, keep longest). Match = Phrase. Volume/CPC = expert estimates for Germany.",8)
header_row(ws,4,["Rank","Score /100","Search pool (English)","Keywords — German (paste as-is, phrase match)","English meaning","Volume","CPC € (est.)","Role & cut order"])
pools=[
 (1,100,"Google Ads agency",
   [("google ads agentur","Google Ads agency"),("adwords agentur","AdWords agency"),
    ("google adwords agentur","Google AdWords agency"),("agentur für google ads","agency for Google Ads"),
    ("google ads dienstleister","Google Ads service provider")],
   "High","€4–13","Anchor — keep always. Spends first, drives most volume."),
 (2,93,"Done-for-me (run / build / set up)",
   [("google ads schalten lassen","have Google Ads run for me"),
    ("google ads kampagne erstellen lassen","have a campaign built"),
    ("google ads einrichten lassen","have Google Ads set up")],
   "Low","€3.50–9","⭐ Highest intent. 'Lassen' [to have done by someone] = explicit outsource signal. Best conversion rate. Keep all."),
 (3,90,"Explicit hire intent",
   [("google ads agentur beauftragen","commission / hire a Google Ads agency"),
    ("google ads outsourcing","Google Ads outsourcing")],
   "Very low","€3–8","⭐ 'Beauftragen' [to commission someone] = bottom-funnel buyer ready to act. Very low volume but hottest intent."),
 (4,87,"Ongoing management",
   [("google ads betreuung","Google Ads ongoing management"),("google ads management","Google Ads management")],
   "Mid","€3.50–9.50","Keep — recurring-revenue hire intent."),
 (5,84,"Consulting",
   [("google ads beratung","Google Ads consulting")],
   "Mid","€3–9","Keep — maps to your Calendly consultation (primary goal)."),
 (6,76,"Solo specialist",
   [("google ads freelancer","Google Ads freelancer"),("google ads experte","Google Ads expert"),("google ads spezialist","Google Ads specialist")],
   "Low","€2.50–8","Keep — fits your solo positioning. Note: 'freelancer' attracts budget-conscious buyers."),
 (7,66,"Optimize / fix my account",
   [("google ads optimieren lassen","have Google Ads optimized"),("google ads optimierung","Google Ads optimization")],
   "Mid","€2.50–7.50","Test — 'optimieren lassen' [have optimized] = hire; 'optimierung' = watch for DIY."),
 (8,58,"Agency cost",
   [("google ads agentur kosten","Google Ads agency cost")],
   "Low","€4.50–10","Test — price shopper, but comparing agencies = a buyer."),
 (9,44,"Aspirational",
   [("mehr kunden mit google ads","more customers with Google Ads")],
   "Low","€2–6","Test — mixed intent. Cut if no conversions after 3 weeks."),
 (10,40,"Pain symptoms",
   [("google ads bringt nichts","Google Ads brings nothing"),("google ads keine conversions","Google Ads no conversions"),("google ads zu teuer","Google Ads too expensive")],
   "Very low","€2–6","Free option — tiny volume, great if it fires. Cut order #2."),
 (11,28,"Platform cost (watch)",
   [("google ads kosten","Google Ads cost")],
   "High","€2–5.50","Watch closely — researchers / DIY. Cut order #1 (pause first)."),
]
def score_color(s):
    if s>=85: return GREEN
    if s>=66: return NAVY
    if s>=44: return AMBER
    return RED
voltxt={"High":GREEN,"Mid":AMBER,"Low":GREY,"Very low":GREY}
r=5
for rank,score,pool,kw,vol,cpc,role in pools:
    fill=LIGHT if rank%2 else WHITE
    de_block="\n".join(k[0] for k in kw); en_block="\n".join(k[1] for k in kw)
    for j,v in enumerate([rank,score,pool,de_block,en_block,vol,cpc,role],1):
        c=ws.cell(r,j,v); c.border=border; c.fill=PatternFill("solid",fgColor=fill)
        if j==1: c.font=cfont(11,bold=True,color=NAVY); c.alignment=center
        elif j==2: c.font=cfont(14,bold=True,color=score_color(score)); c.alignment=center
        elif j==3: c.font=cfont(10,bold=True,color=NAVY); c.alignment=topleft
        elif j==4: c.font=cfont(10,bold=True,color=NAVY); c.alignment=topleft
        elif j==5: c.font=cfont(9,italic=True,color=GREY); c.alignment=topleft
        elif j==6: c.font=cfont(10,bold=True,color=voltxt[vol]); c.alignment=center
        elif j==7: c.font=cfont(10); c.alignment=center
        else: c.font=cfont(9,color=GREY); c.alignment=topleft
    ws.row_dimensions[r].height=max(28,15*len(kw)+14); r+=1
for k,v in {'A':6,'B':10,'C':22,'D':34,'E':28,'F':10,'G':12,'H':36}.items(): ws.column_dimensions[k].width=v
ws.freeze_panes="A5"
total_kw=sum(len(p[3]) for p in pools)
r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
d=ws.cell(r,1,f"{total_kw} keywords → {len(pools)} search pools. You can enter all {total_kw} keywords (Google dedupes within a pool, so synonyms don't compete — no harm). Score = priority: run ranks 1–4 for sure, test 5–9, watch 10–11. If you trim after ~3 weeks, cut from the bottom up. Volume & CPC are expert estimates for Germany (German language), not live Keyword Planner numbers.")
d.font=cfont(9,italic=True,color=GREY); d.alignment=topleft; ws.row_dimensions[r].height=52

# ============================================================ 3. NEGATIVE KEYWORDS
ws=wb.create_sheet("3. Negative Keywords")
title_block(ws,"Negative keywords — add BEFORE launch",
            "Filters out job seekers, learners, free/DIY searchers. Recommended: phrase match at campaign level. German term goes live; English beside it.",4)
header_row(ws,4,["#","Negative keyword (German)","English translation","Why exclude"])
negs=[
 ("kostenlos","free","Free-seekers"),("gratis","free","Free-seekers"),("umsonst","for free","Free-seekers"),
 ("job","job","Job seekers"),("jobs","jobs","Job seekers"),("stellenangebote","job offers","Job seekers"),
 ("gehalt","salary","Salary research"),("ausbildung","apprenticeship","Apprenticeship seekers"),("praktikum","internship","Internship seekers"),
 ("werkstudent","working student","Job seekers"),("weiterbildung","further education","Learners"),("kurs","course","Learners"),
 ("kurse","courses","Learners"),("schulung","training","Learners"),("seminar","seminar","Learners"),
 ("zertifizierung","certification","Learners / exam"),("zertifikat","certificate","Learners / exam"),("tutorial","tutorial","DIY / learners"),
 ("anleitung","guide / how-to","DIY searchers"),("lernen","to learn","Learners"),("selber machen","do it yourself","DIY, no hire intent"),
 ("selbst","yourself","DIY signal"),("udemy","udemy","Course platform"),("ihk","chamber of commerce","Education / training"),
 ("definition","definition","Info search"),("was ist","what is","Info search"),("bedeutung","meaning","Info search"),
 ("erklärung","explanation","Info search"),("login","login","Account / support"),("einloggen","log in","Account search"),
 ("konto","account","Account search"),("gutschein","voucher","Discount seekers"),("guthaben","credit / coupon","Discount seekers"),
 ("billig","cheap","Bargain seekers"),("google ads zertifizierung","Google Ads certification","Learners / exam"),("google ads kurs","Google Ads course","Learners"),
]
r=5
for i,(de,en,why) in enumerate(negs,1):
    fill=LIGHT if i%2 else WHITE
    for j,v in enumerate([i,de,en,why],1):
        c=ws.cell(r,j,v); c.border=border; c.fill=PatternFill("solid",fgColor=fill); c.font=cfont(10)
        if j==1: c.alignment=center
        elif j==2: c.font=cfont(10,bold=True,color=NAVY); c.alignment=left
        elif j==3: c.font=cfont(9,italic=True,color=GREY); c.alignment=left
        else: c.alignment=left; c.font=cfont(9,color=GREY)
    ws.row_dimensions[r].height=20; r+=1
ws.column_dimensions['A'].width=5; ws.column_dimensions['B'].width=30; ws.column_dimensions['C'].width=24; ws.column_dimensions['D'].width=34
ws.freeze_panes="A5"; ws.auto_filter.ref=f"A4:D{r-1}"

# ============================================================ 4. ADS (RSA)
ws=wb.create_sheet("4. Ads (RSA)")
title_block(ws,"Responsive search ads — Set A & Set B (A/B test)",
            "Headline ≤30 chars · Description ≤90 chars (incl. spaces). German goes live; English translation beside it. 'Chars' = verified length.",5)
for k,v in {'A':5,'B':46,'C':46,'D':8,'E':12}.items(): ws.column_dimensions[k].width=v
def asset_block(ws,start,label,items,limit,kind):
    band(ws,start,label,5); hr=start+1
    for j,h in enumerate(["#","Text (German — goes live)","English translation","Chars","Status"],1):
        c=ws.cell(hr,j,h); c.font=hfont(10); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=center; c.border=border
    ws.row_dimensions[hr].height=24; r=hr+1
    for i,(de,en) in enumerate(items,1):
        fill=LIGHT if i%2 else WHITE
        ws.cell(r,1,i).alignment=center; ws.cell(r,1).border=border; ws.cell(r,1).fill=PatternFill("solid",fgColor=fill); ws.cell(r,1).font=cfont(10)
        cde=ws.cell(r,2,de); cde.font=cfont(10,bold=True,color=NAVY); cde.alignment=wrap; cde.border=border; cde.fill=PatternFill("solid",fgColor=fill)
        cen=ws.cell(r,3,en); cen.font=cfont(9,italic=True,color=GREY); cen.alignment=wrap; cen.border=border; cen.fill=PatternFill("solid",fgColor=fill)
        n,status=chk(de,limit,label)
        cz=ws.cell(r,4,n); cz.alignment=center; cz.border=border; cz.fill=PatternFill("solid",fgColor=fill); cz.font=cfont(10)
        cs=ws.cell(r,5,status); cs.alignment=center; cs.border=border; cs.fill=PatternFill("solid",fgColor=fill); cs.font=cfont(10,bold=True,color=GREEN if status=="OK" else RED)
        ws.row_dimensions[r].height=(30 if kind=="d" else 18); r+=1
    return r
band(ws,4,"SET A — Benefit & proof",5)
A_h=[("Google Ads Agentur","Google Ads agency"),("Endlich Kunden statt Klicks","Finally customers, not just clicks"),
 ("Schluss mit Streuverlust","End wasted ad spend"),("Mehr Anfragen, weniger Kosten","More inquiries, lower costs"),
 ("Werbebudget, das sich lohnt","Ad budget that pays off"),("30 Tage testen, kein Risiko","Test 30 days, no risk"),
 ("Kostenlose Erstberatung","Free initial consultation"),("Transparent & datenbasiert","Transparent & data-driven"),
 ("Jeder Euro messbar","Every euro measurable"),("Monatlich kündbar","Cancel monthly"),
 ("Google & Meta Ads Profis","Google & Meta Ads pros"),("Über €1,7 Mio. Umsatz","Over €1.7M revenue generated"),
 ("72 Projekte, 15 Länder","72 projects, 15 countries"),("Ads-Betreuung ab €499/Mo.","Ads management from €499/mo"),
 ("Jetzt Termin sichern","Book your appointment now")]
r=asset_block(ws,5,"Set A · Headlines (max 30)",A_h,30,"h")
A_d=[("Viele Klicks, aber keine Kunden? Wir bauen Google Ads, bei denen jeder Euro arbeitet.","Lots of clicks but no customers? We build Google Ads where every euro works."),
 ("Kostenlose Erstberatung in 30 Minuten. Transparent, monatlich kündbar, ohne Risiko.","Free 30-minute initial consultation. Transparent, cancel monthly, no risk."),
 ("Über €1,7 Mio. Umsatz generiert. Jetzt unverbindlichen Beratungstermin buchen.","Over €1.7M revenue generated. Book a no-obligation consultation now."),
 ("Schluss mit verbranntem Werbebudget. Mehr qualifizierte Anfragen ab Tag eins.","End burned ad budget. More qualified inquiries from day one.")]
r=asset_block(ws,r,"Set A · Descriptions (max 90)",A_d,90,"d"); r+=1
B_h=[("Werbebudget verbrannt?","Ad budget burned?"),("Viele Klicks, keine Kunden?","Lots of clicks, no customers?"),
 ("Google Ads ohne Ergebnis?","Google Ads with no results?"),("Endlich planbare Anfragen","Finally predictable inquiries"),
 ("Wir retten Ihr Budget","We rescue your budget"),("Mehr Leads, weniger Kosten","More leads, lower costs"),
 ("Ehrliches Performance-Ads","Honest performance ads"),("30-Tage-Pilot, kein Risiko","30-day pilot, no risk"),
 ("Kostenlose Analyse","Free analysis"),("Daten statt Bauchgefühl","Data instead of gut feeling"),
 ("Jeder Euro messbar","Every euro measurable"),("Schluss mit Rätselraten","End the guesswork"),
 ("Google & Meta aus 1 Hand","Google & Meta from one source"),("Beratung in 30 Minuten","Consultation in 30 minutes"),
 ("Jetzt Termin buchen","Book an appointment now")]
r=asset_block(ws,r,"SET B — Pain-point & emotion · Headlines (max 30)",B_h,30,"h")
B_d=[("Ihre Google Ads bringen Klicks, aber keine Anfragen? Wir finden das Leck im Tracking.","Your Google Ads bring clicks but no inquiries? We find the leak in your tracking."),
 ("Kostenlose Analyse Ihrer Kampagnen. 30-Tage-Pilot ohne Risiko, monatlich kündbar.","Free analysis of your campaigns. 30-day pilot, no risk, cancel monthly."),
 ("Keine leeren Versprechen, nur messbare Ergebnisse. Jetzt Beratungstermin sichern.","No empty promises, only measurable results. Secure a consultation now."),
 ("Wir holen mehr Kunden aus Ihrem Budget - datengetrieben, transparent, ehrlich.","We get more customers from your budget - data-driven, transparent, honest.")]
r=asset_block(ws,r,"Set B · Descriptions (max 90)",B_d,90,"d"); r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
tip=ws.cell(r,1,"Tip: each RSA allows up to 15 headlines and 4 descriptions — enter all of a set and let Google test combinations. Optionally pin 2–3 (e.g. headline position 1 = brand name).")
tip.font=cfont(9,italic=True,color=GREY); tip.alignment=topleft; ws.row_dimensions[r].height=34

# ============================================================ 5. CALLOUTS & SNIPPETS
ws=wb.create_sheet("5. Callouts & Snippets")
title_block(ws,"Callout extensions & structured snippets",
            "Callout ≤25 chars · snippet value ≤25 chars. German goes live; English beside it. 'Chars' = verified length.",5)
for k,v in {'A':5,'B':40,'C':40,'D':8,'E':12}.items(): ws.column_dimensions[k].width=v
callouts=[("Kostenlose Erstberatung","Free initial consultation"),("30-Tage-Pilotprojekt","30-day pilot project"),
 ("Monatlich kündbar","Cancel monthly"),("Transparente Reports","Transparent reports"),("Google & Meta Ads","Google & Meta Ads"),
 ("DSGVO-konform","GDPR-compliant"),("Antwort in 24 Stunden","Reply within 24 hours"),("Kein Jahresvertrag","No annual contract"),
 ("Persönliche Betreuung","Personal support"),("Über €1,7 Mio. Umsatz","Over €1.7M revenue")]
def simple_assets(ws,start,label,items,limit):
    band(ws,start,label,5); hr=start+1
    for j,h in enumerate(["#","Text (German — goes live)","English translation","Chars","Status"],1):
        c=ws.cell(hr,j,h); c.font=hfont(10); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=center; c.border=border
    ws.row_dimensions[hr].height=24; r=hr+1
    for i,(de,en) in enumerate(items,1):
        fill=LIGHT if i%2 else WHITE
        ws.cell(r,1,i).alignment=center; ws.cell(r,1).border=border; ws.cell(r,1).fill=PatternFill("solid",fgColor=fill); ws.cell(r,1).font=cfont(10)
        cde=ws.cell(r,2,de); cde.font=cfont(10,bold=True,color=NAVY); cde.alignment=left; cde.border=border; cde.fill=PatternFill("solid",fgColor=fill)
        cen=ws.cell(r,3,en); cen.font=cfont(9,italic=True,color=GREY); cen.alignment=left; cen.border=border; cen.fill=PatternFill("solid",fgColor=fill)
        n,status=chk(de,limit,label)
        cz=ws.cell(r,4,n); cz.alignment=center; cz.border=border; cz.fill=PatternFill("solid",fgColor=fill); cz.font=cfont(10)
        cs=ws.cell(r,5,status); cs.alignment=center; cs.border=border; cs.fill=PatternFill("solid",fgColor=fill); cs.font=cfont(10,bold=True,color=GREEN if status=="OK" else RED)
        ws.row_dimensions[r].height=18; r+=1
    return r
r=simple_assets(ws,4,"Callout extensions (max 25)",callouts,25); r+=1
band(ws,r,"Structured snippets — create 2 (set snippet language = German, then pick the Header from Google's fixed dropdown)",5); r+=1
for j,h in enumerate(["Header (pick in Google Ads)","Values (German — go live)","Values (English)","Note"],1):
    c=ws.cell(r,j,h); c.font=hfont(10); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=center; c.border=border
ws.merge_cells(start_row=r,start_column=4,end_row=r,end_column=5)
ws.row_dimensions[r].height=24; r+=1
snips=[
 ("Serviceleistungen  (= Service catalog)","Google Ads\nMeta Ads\nSuchmaschinenwerbung\nYouTube Ads\nTracking & Analytics","Google Ads\nMeta Ads\nSearch advertising\nYouTube Ads\nTracking & Analytics","Enter each line as a separate value"),
 ("Typen  (= Types)","E-Commerce\nLead-Generierung\nB2B-Marketing\nSaaS\nOnlineshops","E-Commerce\nLead generation\nB2B marketing\nSaaS\nOnline shops","Enter each line as a separate value"),
]
for i,(hd,vd,ve,note) in enumerate(snips,1):
    fill=LIGHT if i%2 else WHITE
    for j,v in enumerate([hd,vd,ve,note],1):
        c=ws.cell(r,j,v); c.border=border; c.fill=PatternFill("solid",fgColor=fill)
        if j==1: c.font=cfont(10,bold=True,color=NAVY); c.alignment=topleft
        elif j==2: c.font=cfont(10); c.alignment=topleft
        elif j==3: c.font=cfont(9,italic=True,color=GREY); c.alignment=topleft
        else: c.font=cfont(9,color=GREY); c.alignment=topleft
    ws.cell(r,5).border=border; ws.cell(r,5).fill=PatternFill("solid",fgColor=fill)
    ws.row_dimensions[r].height=82; r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
nt=ws.cell(r,1,"Each value is listed on its own line — in Google, add them one at a time (don't type bullets, dots or commas between them). Only these 2 headers fit your offer; the header list is fixed by Google, so 'Highlights' is not an option. USP phrases (Monatlich kündbar, DSGVO-konform) belong in your Callouts above.")
nt.font=cfont(9,italic=True,color=GREY); nt.alignment=topleft; ws.row_dimensions[r].height=40; r+=1
ws.column_dimensions['A'].width=32; ws.column_dimensions['B'].width=40; ws.column_dimensions['C'].width=40; ws.column_dimensions['D'].width=22; ws.column_dimensions['E'].width=4

# ============================================================ 6. SITELINKS
ws=wb.create_sheet("6. Sitelinks")
title_block(ws,"Sitelink extensions",
            "Title ≤25 chars · each description line ≤35 chars. German goes live; English translation beside each. 'Chars' = title / line1 / line2.",9)
header_row(ws,4,["#","Title (German)","Title (English)","Description 1 (German)","Description 1 (English)","Description 2 (German)","Description 2 (English)","Target URL","Chars"])
sitelinks=[
 ("Kostenlose Erstberatung","Free consultation","In 30 Minuten zu mehr Anfragen","To more inquiries in 30 minutes","Unverbindlich & kostenlos","No obligation & free","/#contact"),
 ("30-Tage-Pilotprojekt","30-day pilot","Testen Sie uns ohne Risiko","Test us with no risk","Monatlich kündbar","Cancel monthly","/#pilot"),
 ("Unsere Leistungen","Our services","Google, Meta & YouTube Ads","Google, Meta & YouTube Ads","Alles aus einer Hand","Everything from one source","/#services"),
 ("Unsere Preise","Our pricing","Transparente Pakete ab €499","Transparent packages from €499","Kein Jahresvertrag","No annual contract","/#pricing"),
 ("So arbeiten wir","How we work","Unser Prozess in 6 Schritten","Our 6-step process","Transparent & datenbasiert","Transparent & data-driven","/#process"),
 ("Erfolge ansehen","See our results","Echte Cases & Ergebnisse","Real cases & results","Über €1,7 Mio. Umsatz","Over €1.7M revenue","/#cases"),
]
r=5
for i,(t,te,d1,d1e,d2,d2e,url) in enumerate(sitelinks,1):
    fill=LIGHT if i%2 else WHITE
    chk(t,25,"Sitelink title"); chk(d1,35,"Sitelink line1"); chk(d2,35,"Sitelink line2")
    vals=[i,t,te,d1,d1e,d2,d2e,url,f"{len(t)}/{len(d1)}/{len(d2)}"]
    for j,v in enumerate(vals,1):
        c=ws.cell(r,j,v); c.border=border; c.fill=PatternFill("solid",fgColor=fill); c.font=cfont(10)
        if j==1 or j==9: c.alignment=center
        elif j in (2,4,6): c.font=cfont(10,bold=True,color=NAVY); c.alignment=wrap
        elif j in (3,5,7): c.font=cfont(9,italic=True,color=GREY); c.alignment=wrap
        else: c.font=cfont(9,color=GREY); c.alignment=left
    ws.row_dimensions[r].height=30; r+=1
for k,v in {'A':5,'B':24,'C':20,'D':28,'E':28,'F':22,'G':24,'H':14,'I':12}.items(): ws.column_dimensions[k].width=v
ws.freeze_panes="A5"
r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=9)
n=ws.cell(r,1,"Limits: title ≤25, each description line ≤35 chars. Activate at least 4 sitelinks (6 is better). Anchor URLs assume your one-page site sections.")
n.font=cfont(9,italic=True,color=GREY); n.alignment=topleft; ws.row_dimensions[r].height=28

# ============================================================ 7. AUDIENCES
ws=wb.create_sheet("7. Audiences")
title_block(ws,"Audiences — add in OBSERVATION mode (English, as in your Google Ads UI)",
            "On $30/day do NOT use 'Targeting' (it cuts reach). 'Observation' just watches who responds. Names below are exactly as they appear in an English Google Ads account.",4)
header_row(ws,4,["Audience (name in Google Ads)","What to set up","Mode","Why / how to use"])
auds=[
 ("Custom segment (search intent)",
  "Create a custom segment → pick 'People who searched for any of these terms on Google', then paste these German search terms: \"google ads agentur\" (Google Ads agency), \"online marketing agentur\" (online marketing agency), \"performance marketing agentur\" (performance marketing agency), \"sea agentur\" (SEA agency) — plus the names of competitor agencies you know.",
  "Observation","Your hottest audience: people actively searching for what you sell. On Search, a custom segment uses SEARCH TERMS only — this is also how you reach compare-shoppers (add competitor names here)."),
 ("In-market: Advertising & Marketing Services",
  "Add this segment from Google's in-market list.","Observation","People currently in the market for marketing services."),
 ("In-market: Business Services / Business Technology",
  "Add this segment from Google's in-market list.","Observation","B2B decision-makers, software/tech context."),
 ("Your data: Website visitors (remarketing)",
  "Auto-built from your GTM/GA4 tag — past site visitors (last 30–90 days).","Observation","Warm audience; only usable once your tag has collected enough visitors."),
 ("Your data: Customer Match",
  "Upload your existing client / lead email list.","Observation","Lets Google find similar people; you can also exclude current clients."),
 ("Detailed demographics: Industry / company size",
  "Add if shown for your account.","Observation","Fine-tuning only; coverage is limited in Germany."),
]
r=5
for i,(t,what,mode,why) in enumerate(auds,1):
    fill=LIGHT if i%2 else WHITE
    for j,v in enumerate([t,what,mode,why],1):
        c=ws.cell(r,j,v); c.border=border; c.fill=PatternFill("solid",fgColor=fill)
        if j==1: c.font=cfont(10,bold=True,color=NAVY); c.alignment=topleft
        elif j==2: c.font=cfont(10); c.alignment=topleft
        elif j==3: c.font=cfont(10,bold=True,color=RED); c.alignment=center
        else: c.font=cfont(9,color=GREY); c.alignment=topleft
    ws.row_dimensions[r].height=58; r+=1
for k,v in {'A':36,'B':58,'C':13,'D':40}.items(): ws.column_dimensions[k].width=v
ws.freeze_panes="A5"
band(ws,r,"Available on Search but SKIP for your offer: 'Affinity segments' (broad lifestyle interests) and 'Life events' (marriage, moving, graduation) — not relevant to a B2B agency.",4); r+=1
band(ws,r,"HOW TO USE",4); r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
how=ws.cell(r,1,"1) Add all segments as Observation — your reach stays full. 2) After 2–3 weeks, open the Audiences tab inside Google Ads to see which segment produced conversions. 3) Increase the bid on winners (+10–20%); lower the bid on, or exclude, weak segments. The German terms in row 1 are quoted exactly as you paste them into the custom segment (English meaning is in brackets).")
how.font=cfont(10); how.alignment=topleft; ws.row_dimensions[r].height=64

# ---- validate ----
if OVER:
    for where,text,n,limit in OVER: print(f"OVER [{where}] ({n}/{limit}): {text}")
    raise SystemExit("Build aborted: asset over limit.")
print("All assets within character limits.")
import shutil
out1="/Users/mdanyalshahzad/Documents/Danyal/Freelancing/Digital Marketing Remote/Website/DMR-Google-Ads-Kampagne-Komplett-DE.xlsx"
out2="/Users/mdanyalshahzad/Downloads/DMR-Google-Ads-Kampagne-Komplett-DE.xlsx"
wb.save(out1); shutil.copyfile(out1,out2)
print("saved",out1); print("saved",out2)
